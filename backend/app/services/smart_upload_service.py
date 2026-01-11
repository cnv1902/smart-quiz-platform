"""
Smart Upload Service - AI-Powered Document Classification and S3 Organization

This service implements the "Smart Folder" feature that:
1. Extracts text from uploaded documents (first 5 questions)
2. Uses Gemini AI to classify the document's category and subject
3. Deduplicates topics using fuzzy matching against existing DB records
4. Uploads files to S3 with dynamic path: {category_slug}/{topic_slug}/{filename}

Example Flow:
    Input: "DeThi_KinhTeViMo.pdf"
    -> Extract text: "Câu 1: Cầu về hàng hóa là gì?..."
    -> Gemini Response: {"category": "Kinh Tế", "subject": "Kinh Tế Vi Mô"}
    -> S3 Key: "kinh-te/kinh-te-vi-mo/dethi-kinhtevimu.pdf"
"""

import json
import logging
import re
import time
import unicodedata
from io import BytesIO
from typing import Optional, Tuple

import boto3
from botocore.exceptions import ClientError, NoCredentialsError
from docx import Document
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from thefuzz import process

try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

from app.core.config import settings
from app.models.knowledge_base import Topic, Resource

logger = logging.getLogger(__name__)


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def slugify(text: str) -> str:
    """
    Convert Vietnamese text to URL-safe slug.
    
    Examples:
        "Kinh Tế Vi Mô" -> "kinh-te-vi-mo"
        "Toán Cao Cấp" -> "toan-cao-cap"
    """
    vietnamese_map = {
        'à': 'a', 'á': 'a', 'ả': 'a', 'ã': 'a', 'ạ': 'a',
        'ă': 'a', 'ằ': 'a', 'ắ': 'a', 'ẳ': 'a', 'ẵ': 'a', 'ặ': 'a',
        'â': 'a', 'ầ': 'a', 'ấ': 'a', 'ẩ': 'a', 'ẫ': 'a', 'ậ': 'a',
        'đ': 'd',
        'è': 'e', 'é': 'e', 'ẻ': 'e', 'ẽ': 'e', 'ẹ': 'e',
        'ê': 'e', 'ề': 'e', 'ế': 'e', 'ể': 'e', 'ễ': 'e', 'ệ': 'e',
        'ì': 'i', 'í': 'i', 'ỉ': 'i', 'ĩ': 'i', 'ị': 'i',
        'ò': 'o', 'ó': 'o', 'ỏ': 'o', 'õ': 'o', 'ọ': 'o',
        'ô': 'o', 'ồ': 'o', 'ố': 'o', 'ổ': 'o', 'ỗ': 'o', 'ộ': 'o',
        'ơ': 'o', 'ờ': 'o', 'ớ': 'o', 'ở': 'o', 'ỡ': 'o', 'ợ': 'o',
        'ù': 'u', 'ú': 'u', 'ủ': 'u', 'ũ': 'u', 'ụ': 'u',
        'ư': 'u', 'ừ': 'u', 'ứ': 'u', 'ử': 'u', 'ữ': 'u', 'ự': 'u',
        'ỳ': 'y', 'ý': 'y', 'ỷ': 'y', 'ỹ': 'y', 'ỵ': 'y',
    }
    
    text = text.lower()
    result = []
    for char in text:
        if char in vietnamese_map:
            result.append(vietnamese_map[char])
        elif char.upper() in vietnamese_map:
            result.append(vietnamese_map[char.upper()])
        else:
            result.append(char)
    text = ''.join(result)
    
    text = unicodedata.normalize('NFKD', text)
    text = text.encode('ascii', 'ignore').decode('ascii')
    text = re.sub(r'[^\w\s.-]', '', text)
    text = re.sub(r'[\s_]+', '-', text)
    text = re.sub(r'-+', '-', text)
    text = text.strip('-')
    
    return text


def sanitize_filename(filename: str) -> str:
    """
    Sanitize filename for S3 storage.
    Example: "Đề Thi 2024.pdf" -> "de-thi-2024.pdf"
    """
    if '.' in filename:
        name, ext = filename.rsplit('.', 1)
        return f"{slugify(name)}.{ext.lower()}"
    return slugify(filename)


# =============================================================================
# TEXT EXTRACTION
# =============================================================================

def extract_text_from_docx(file_content: bytes, max_questions: int = 5) -> str:
    """
    Extract text from Word document, focusing on first N questions.
    """
    try:
        doc = Document(BytesIO(file_content))
        paragraphs = []
        question_count = 0
        question_pattern = re.compile(r'^(?:Câu|Question|Q)\s*\d+', re.IGNORECASE)
        
        for para in doc.paragraphs:
            text = para.text.strip()
            if not text:
                continue
            
            if question_pattern.match(text):
                question_count += 1
                if question_count > max_questions:
                    break
            
            paragraphs.append(text)
        
        return '\n'.join(paragraphs)
    except Exception as e:
        logger.warning(f"Failed to extract text from DOCX: {e}")
        return ""


def extract_text_from_xlsx(file_content: bytes, max_rows: int = 20) -> str:
    """
    Extract text from Excel file.
    """
    try:
        wb = load_workbook(BytesIO(file_content), read_only=True)
        texts = []
        
        for sheet in wb.worksheets:
            for row_idx, row in enumerate(sheet.iter_rows(max_row=max_rows, values_only=True)):
                if row_idx >= max_rows:
                    break
                row_text = ' '.join(str(cell) for cell in row if cell)
                if row_text.strip():
                    texts.append(row_text)
        
        return '\n'.join(texts)
    except Exception as e:
        logger.warning(f"Failed to extract text from XLSX: {e}")
        return ""


def extract_text_from_txt(file_content: bytes, max_chars: int = 3000) -> str:
    """
    Extract text from plain text file.
    """
    try:
        text = file_content.decode('utf-8', errors='ignore')
        return text[:max_chars]
    except Exception as e:
        logger.warning(f"Failed to extract text from TXT: {e}")
        return ""


def extract_text(file_content: bytes, filename: str, max_questions: int = 5) -> str:
    """
    Extract text from file based on extension.
    """
    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    
    if ext == 'docx':
        return extract_text_from_docx(file_content, max_questions)
    elif ext in ('xlsx', 'xls'):
        return extract_text_from_xlsx(file_content)
    elif ext == 'txt':
        return extract_text_from_txt(file_content)
    else:
        # For PDF and other formats, try as text
        return extract_text_from_txt(file_content)


# =============================================================================
# SMART UPLOAD SERVICE
# =============================================================================

class SmartUploadService:
    """
    Smart Upload Service that uses AI to classify documents and organize them
    into the correct S3 folder structure.
    
    S3 Structure: {category_slug}/{topic_slug}/{filename}
    Example: kinh-te/kinh-te-vi-mo/de-thi-2024.pdf
    """
    
    # Default classification for fallback
    DEFAULT_CATEGORY = "Tài liệu chung"
    DEFAULT_SUBJECT = "Tài liệu"
    
    def __init__(self):
        """Initialize service with S3 client and Gemini AI."""
        # Initialize S3
        self.bucket_name = settings.aws_bucket_name
        self.region = settings.aws_region
        
        if not self.bucket_name:
            raise ValueError("AWS_BUCKET_NAME is required.")
        
        self.s3_client = boto3.client(
            's3',
            aws_access_key_id=settings.aws_access_key_id,
            aws_secret_access_key=settings.aws_secret_access_key,
            region_name=self.region
        )
        
        # Initialize Gemini
        self.gemini_model = None
        if GEMINI_AVAILABLE and settings.gemini_api_key:
            try:
                genai.configure(api_key=settings.gemini_api_key)
                self.gemini_model = genai.GenerativeModel('gemini-2.5-flash')
                logger.info("Gemini AI initialized successfully")
            except Exception as e:
                logger.warning(f"Failed to initialize Gemini: {e}")
        else:
            logger.warning("Gemini AI not available. Will use default classification.")
        
        logger.info(f"SmartUploadService initialized for bucket: {self.bucket_name}")

    # =========================================================================
    # STEP 1: ANALYZE - Extract text and classify with AI
    # =========================================================================
    
    def _analyze_document(self, file_content: bytes, filename: str) -> dict:
        """
        Step 1: Analyze document content using Gemini AI.
        
        Returns:
            {"category": "Kinh Tế", "subject": "Kinh Tế Vi Mô"}
        """
        # Extract text from document
        extracted_text = extract_text(file_content, filename, max_questions=5)
        
        if not extracted_text or len(extracted_text.strip()) < 50:
            logger.warning(f"Not enough text extracted from {filename}, using filename for classification")
            extracted_text = filename
        
        # If Gemini not available, use fallback
        if not self.gemini_model:
            return self._fallback_classification(filename, extracted_text)
        
        # Call Gemini API
        try:
            prompt = self._build_classification_prompt(extracted_text, filename)
            response = self.gemini_model.generate_content(prompt)
            result = self._parse_gemini_response(response.text)
            
            logger.info(f"Gemini classified '{filename}' as: {result}")
            return result
            
        except Exception as e:
            logger.error(f"Gemini API error: {e}")
            return self._fallback_classification(filename, extracted_text)
    
    def _build_classification_prompt(self, text: str, filename: str) -> str:
        """Build prompt for Gemini classification."""
        return f"""Phân tích nội dung tài liệu sau và xác định danh mục (category) và môn học (subject).

Tên file: {filename}

Nội dung:
{text[:2000]}

Hãy trả về CHÍNH XÁC theo format JSON sau (không thêm bất kỳ text nào khác):
{{"category": "Tên danh mục", "subject": "Tên môn học"}}

Ví dụ các category phổ biến:
- Kinh Tế (cho các môn: Kinh Tế Vi Mô, Kinh Tế Vĩ Mô, Tài Chính, Kế Toán...)
- Toán học (cho các môn: Giải Tích, Đại Số, Xác Suất Thống Kê...)
- Công nghệ (cho các môn: Lập Trình, Cơ Sở Dữ Liệu, Mạng Máy Tính...)
- Ngoại ngữ (cho các môn: Tiếng Anh, Tiếng Nhật...)
- Khoa học (cho các môn: Vật Lý, Hóa Học, Sinh Học...)
- Đề thi (cho các đề thi tổng hợp không rõ môn)

Chỉ trả về JSON, không giải thích."""

    def _parse_gemini_response(self, response_text: str) -> dict:
        """Parse Gemini response to extract category and subject."""
        try:
            # Try to find JSON in response
            json_match = re.search(r'\{[^}]+\}', response_text)
            if json_match:
                result = json.loads(json_match.group())
                category = result.get('category', self.DEFAULT_CATEGORY)
                subject = result.get('subject', self.DEFAULT_SUBJECT)
                return {"category": category, "subject": subject}
        except json.JSONDecodeError:
            pass
        
        return {"category": self.DEFAULT_CATEGORY, "subject": self.DEFAULT_SUBJECT}
    
    def _fallback_classification(self, filename: str, text: str) -> dict:
        """Fallback classification when AI is not available."""
        # Simple keyword-based classification
        combined = f"{filename} {text}".lower()
        
        classifications = [
            (["kinh tế", "kinh te", "economics", "tài chính", "kế toán"], "Kinh Tế", "Kinh Tế"),
            (["toán", "toan", "math", "giải tích", "đại số"], "Toán học", "Toán học"),
            (["lập trình", "lap trinh", "code", "python", "java", "web"], "Công nghệ", "Lập Trình"),
            (["vật lý", "vat ly", "physics"], "Khoa học", "Vật Lý"),
            (["hóa học", "hoa hoc", "chemistry"], "Khoa học", "Hóa Học"),
            (["tiếng anh", "english"], "Ngoại ngữ", "Tiếng Anh"),
            (["đề thi", "de thi", "exam", "kiểm tra"], "Đề thi", "Đề thi"),
        ]
        
        for keywords, category, subject in classifications:
            if any(kw in combined for kw in keywords):
                return {"category": category, "subject": subject}
        
        return {"category": self.DEFAULT_CATEGORY, "subject": self.DEFAULT_SUBJECT}

    # =========================================================================
    # STEP 2: TOPIC DEDUPLICATION - Find or create topic with fuzzy matching
    # =========================================================================
    
    def _find_or_create_topic(
        self, 
        db: Session, 
        subject_name: str, 
        category_name: str
    ) -> Topic:
        """
        Step 2: Find existing topic by exact slug or fuzzy match, or create new.
        
        Deduplication Logic:
        1. Try exact slug match
        2. If not found, fuzzy match against all topic names (threshold > 90%)
        3. If still not found, create new topic
        
        Args:
            db: Database session
            subject_name: Subject name from AI (e.g., "Kinh Tế Vi Mô")
            category_name: Category name from AI (e.g., "Kinh Tế")
            
        Returns:
            Topic object (existing or newly created)
        """
        subject_slug = slugify(subject_name)
        category_slug = slugify(category_name)
        
        logger.info(f"Finding topic for subject='{subject_name}' (slug={subject_slug}), category='{category_name}'")
        
        # Step 2a: Try exact slug match
        existing = db.query(Topic).filter(Topic.slug == subject_slug).first()
        if existing:
            logger.info(f"Found existing topic by exact slug: {existing.name}")
            # Update category if missing
            if not existing.category_slug:
                existing.category = category_name
                existing.category_slug = category_slug
                db.commit()
            return existing
        
        # Step 2b: Fuzzy match against all topic names
        all_topics = db.query(Topic).all()
        if all_topics:
            topic_names = [t.name for t in all_topics]
            match_result = process.extractOne(subject_name, topic_names, score_cutoff=90)
            
            if match_result:
                matched_name, score = match_result[0], match_result[1]
                matched_topic = next(t for t in all_topics if t.name == matched_name)
                logger.info(f"Fuzzy matched '{subject_name}' to '{matched_name}' (score={score})")
                
                # Update category if missing
                if not matched_topic.category_slug:
                    matched_topic.category = category_name
                    matched_topic.category_slug = category_slug
                    db.commit()
                
                return matched_topic
        
        # Step 2c: No match found - create new topic
        new_topic = Topic(
            name=subject_name,
            slug=subject_slug,
            category=category_name,
            category_slug=category_slug,
            description=f"Môn học: {subject_name} - Danh mục: {category_name}"
        )
        db.add(new_topic)
        db.commit()
        db.refresh(new_topic)
        
        logger.info(f"Created new topic: {new_topic.name} ({new_topic.slug}) in category {category_name}")
        return new_topic

    # =========================================================================
    # STEP 3: BUILD S3 KEY - Dynamic path construction
    # =========================================================================
    
    def _construct_s3_key(self, topic: Topic, filename: str) -> str:
        """
        Step 3: Build S3 key using dynamic path from topic.
        
        CRITICAL: Never use hard-coded paths!
        
        Formula: {category_slug}/{topic_slug}/{sanitized_filename}
        Example: kinh-te/kinh-te-vi-mo/de-thi-2024.pdf
        """
        sanitized = sanitize_filename(filename)
        
        # Build path: category/topic/filename
        if topic.category_slug:
            s3_key = f"{topic.category_slug}/{topic.slug}/{sanitized}"
        else:
            # Fallback: just topic/filename
            s3_key = f"{topic.slug}/{sanitized}"
        
        logger.info(f"Constructed S3 key: {s3_key}")
        return s3_key

    # =========================================================================
    # STEP 4: UPLOAD TO S3
    # =========================================================================
    
    def _upload_to_s3(
        self, 
        file_content: bytes, 
        s3_key: str, 
        content_type: str
    ) -> str:
        """
        Step 4: Upload file to S3 with public-read ACL.
        
        Returns:
            Public URL of uploaded file
        """
        self.s3_client.put_object(
            Bucket=self.bucket_name,
            Key=s3_key,
            Body=file_content,
            ContentType=content_type,
            ACL='public-read'
        )
        
        s3_url = f"https://{self.bucket_name}.s3.{self.region}.amazonaws.com/{s3_key}"
        logger.info(f"File uploaded to: {s3_url}")
        return s3_url

    # =========================================================================
    # MAIN METHOD: SMART UPLOAD
    # =========================================================================
    
    def smart_upload(
        self,
        db: Session,
        file_content: bytes,
        filename: str,
        content_type: str = 'application/octet-stream',
        uploaded_by: Optional[str] = None
    ) -> dict:
        """
        Main method: Smart upload with AI classification.
        
        Complete workflow:
        1. Analyze document with AI -> get category/subject
        2. Find/create topic (with deduplication)
        3. Build dynamic S3 key
        4. Upload to S3
        5. Create database record
        
        Args:
            db: Database session
            file_content: File bytes
            filename: Original filename
            content_type: MIME type
            uploaded_by: User ID
            
        Returns:
            dict with topic, resource, s3_key, s3_url, classification
        """
        logger.info(f"=== SMART UPLOAD START: {filename} ===")
        
        try:
            # Step 1: Analyze with AI
            classification = self._analyze_document(file_content, filename)
            category_name = classification['category']
            subject_name = classification['subject']
            
            logger.info(f"Classification result: category='{category_name}', subject='{subject_name}'")
            
            # Step 2: Find or create topic (with fuzzy deduplication)
            topic = self._find_or_create_topic(db, subject_name, category_name)
            
            # Step 3: Build S3 key dynamically
            s3_key = self._construct_s3_key(topic, filename)
            
            # Handle duplicate filenames
            existing_resource = db.query(Resource).filter(Resource.s3_key == s3_key).first()
            if existing_resource:
                timestamp = int(time.time())
                name_part, ext = s3_key.rsplit('.', 1) if '.' in s3_key else (s3_key, '')
                s3_key = f"{name_part}-{timestamp}.{ext}" if ext else f"{name_part}-{timestamp}"
                logger.info(f"Duplicate detected, new key: {s3_key}")
            
            # Step 4: Upload to S3
            s3_url = self._upload_to_s3(file_content, s3_key, content_type)
            
            # Step 5: Create Resource record
            resource = Resource(
                topic_id=topic.id,
                file_name=filename,
                s3_key=s3_key,
                s3_url=s3_url,
                content_type=content_type,
                file_size=len(file_content),
                uploaded_by=uploaded_by
            )
            db.add(resource)
            db.commit()
            db.refresh(resource)
            
            logger.info(f"=== SMART UPLOAD COMPLETE: {s3_key} ===")
            
            return {
                'classification': classification,
                'topic': {
                    'id': topic.id,
                    'name': topic.name,
                    'slug': topic.slug,
                    'category': topic.category,
                    'category_slug': topic.category_slug
                },
                'resource': {
                    'id': resource.id,
                    'file_name': resource.file_name,
                    's3_key': resource.s3_key,
                    's3_url': resource.s3_url
                },
                's3_key': s3_key,
                's3_url': s3_url,
                'file_id': resource.id,
                'web_view_link': s3_url
            }
            
        except NoCredentialsError:
            logger.error("AWS credentials not found")
            raise Exception("AWS credentials not configured")
        except ClientError as e:
            logger.error(f"S3 error: {e}")
            raise Exception(f"S3 upload failed: {str(e)}")
        except Exception as e:
            logger.error(f"Smart upload failed: {e}")
            raise

    # =========================================================================
    # MANUAL UPLOAD (bypass AI, specify topic directly)
    # =========================================================================
    
    def manual_upload(
        self,
        db: Session,
        file_content: bytes,
        filename: str,
        topic_name: str,
        category_name: Optional[str] = None,
        content_type: str = 'application/octet-stream',
        uploaded_by: Optional[str] = None
    ) -> dict:
        """
        Manual upload without AI classification.
        User specifies topic name directly.
        """
        logger.info(f"=== MANUAL UPLOAD: {filename} to {topic_name} ===")
        
        if not category_name:
            category_name = self.DEFAULT_CATEGORY
        
        # Find or create topic
        topic = self._find_or_create_topic(db, topic_name, category_name)
        
        # Build S3 key
        s3_key = self._construct_s3_key(topic, filename)
        
        # Handle duplicates
        existing = db.query(Resource).filter(Resource.s3_key == s3_key).first()
        if existing:
            timestamp = int(time.time())
            name_part, ext = s3_key.rsplit('.', 1) if '.' in s3_key else (s3_key, '')
            s3_key = f"{name_part}-{timestamp}.{ext}" if ext else f"{name_part}-{timestamp}"
        
        # Upload
        s3_url = self._upload_to_s3(file_content, s3_key, content_type)
        
        # Save to DB
        resource = Resource(
            topic_id=topic.id,
            file_name=filename,
            s3_key=s3_key,
            s3_url=s3_url,
            content_type=content_type,
            file_size=len(file_content),
            uploaded_by=uploaded_by
        )
        db.add(resource)
        db.commit()
        db.refresh(resource)
        
        return {
            'topic': {
                'id': topic.id,
                'name': topic.name,
                'slug': topic.slug,
                'category': topic.category,
                'category_slug': topic.category_slug
            },
            'resource': {
                'id': resource.id,
                'file_name': resource.file_name,
                's3_key': resource.s3_key,
                's3_url': resource.s3_url
            },
            's3_key': s3_key,
            's3_url': s3_url
        }


# =============================================================================
# SERVICE INSTANCE
# =============================================================================

_smart_upload_service: Optional[SmartUploadService] = None


def get_smart_upload_service() -> SmartUploadService:
    """Get or create SmartUploadService singleton."""
    global _smart_upload_service
    if _smart_upload_service is None:
        _smart_upload_service = SmartUploadService()
    return _smart_upload_service
