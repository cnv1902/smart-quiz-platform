"""
Document Parser Service - The "Magic" Part
Parses Word (.docx) and Excel files to extract quiz questions
"""
import re
import uuid
from typing import List, Optional, Tuple
from docx import Document
from docx.shared import RGBColor
from docx.text.paragraph import Paragraph
from docx.text.run import Run
from openpyxl import load_workbook
from io import BytesIO


class QuestionParser:
    """
    Parser for extracting questions from Word and Excel documents
    
    Supports:
    - Question format: "Câu [number]:" or "Question [number]:"
    - Options format: "a.", "b.", "c.", "d." (case insensitive)
    - Bold text detection for correct answers
    """
    
    # Regex patterns
    QUESTION_PATTERN = re.compile(
        r'^(?:Câu|Question|Q)\s*(\d+)\s*[:.]?\s*(.+)',
        re.IGNORECASE | re.MULTILINE
    )
    
    OPTION_PATTERN = re.compile(
        r'^([a-dA-D])\s*[.)\]]\s*(.+)',
        re.IGNORECASE
    )
    
    def __init__(self):
        self.questions = []
    
    def parse_docx(self, file_content: bytes) -> List[dict]:
        """
        Parse a Word document and extract questions
        
        Args:
            file_content: Binary content of the .docx file
            
        Returns:
            List of question dictionaries with options and correct answer
        """
        try:
            doc = Document(BytesIO(file_content))
            return self._extract_questions_from_docx(doc)
        except Exception as e:
            raise ValueError(f"Failed to parse Word document: {str(e)}")
    
    def _extract_questions_from_docx(self, doc: Document) -> List[dict]:
        """Extract questions from parsed Word document"""
        questions = []
        current_question = None
        current_options = []
        
        for paragraph in doc.paragraphs:
            text = paragraph.text.strip()
            if not text:
                continue
            
            # Check if this is a question
            question_match = self.QUESTION_PATTERN.match(text)
            if question_match:
                # Save previous question if exists
                if current_question and current_options:
                    questions.append(self._create_question(
                        current_question, current_options
                    ))
                
                # Start new question
                current_question = {
                    'number': question_match.group(1),
                    'text': question_match.group(2).strip() or self._get_full_text(paragraph),
                    'paragraph': paragraph
                }
                current_options = []
                continue
            
            # Check if this is an option
            option_match = self.OPTION_PATTERN.match(text)
            if option_match and current_question:
                option_letter = option_match.group(1).upper()
                option_text = option_match.group(2).strip()
                
                # Check if this option contains bold text (correct answer)
                is_bold = self._check_bold_in_paragraph(paragraph)
                
                current_options.append({
                    'letter': option_letter,
                    'text': option_text,
                    'is_correct': is_bold
                })
        
        # Don't forget the last question
        if current_question and current_options:
            questions.append(self._create_question(current_question, current_options))
        
        return questions
    
    def _check_bold_in_paragraph(self, paragraph: Paragraph) -> bool:
        """
        Check if a paragraph contains bold text
        This is the CRITICAL feature for auto-detecting correct answers
        """
        for run in paragraph.runs:
            if run.bold:
                return True
        return False
    
    def _get_full_text(self, paragraph: Paragraph) -> str:
        """Get full text from paragraph including all runs"""
        return ''.join(run.text for run in paragraph.runs)
    
    def _create_question(self, question_data: dict, options: List[dict]) -> dict:
        """Create a formatted question dictionary"""
        question_id = str(uuid.uuid4())[:8]
        
        formatted_options = []
        for i, opt in enumerate(options):
            formatted_options.append({
                'id': f"{question_id}_opt_{i}",
                'text': opt['text'],
                'is_correct': opt['is_correct']
            })
        
        # If no option is marked as correct, mark the first one (fallback)
        has_correct = any(opt['is_correct'] for opt in formatted_options)
        if not has_correct and formatted_options:
            # Don't auto-select, let user choose
            pass
        
        return {
            'id': question_id,
            'text': question_data['text'],
            'options': formatted_options
        }
    
    def parse_excel(self, file_content: bytes) -> List[dict]:
        """
        Parse an Excel file and extract questions
        
        Expected format:
        Column A: Question text
        Column B: Option A
        Column C: Option B
        Column D: Option C
        Column E: Option D
        Column F: Correct answer (A/B/C/D)
        
        Args:
            file_content: Binary content of the .xlsx file
            
        Returns:
            List of question dictionaries
        """
        try:
            wb = load_workbook(BytesIO(file_content), read_only=True)
            ws = wb.active
            
            questions = []
            
            # Skip header row
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0]:  # Skip empty rows
                    continue
                
                question_text = str(row[0]).strip()
                if not question_text:
                    continue
                
                question_id = str(uuid.uuid4())[:8]
                
                # Get options (columns B, C, D, E)
                options = []
                option_letters = ['A', 'B', 'C', 'D']
                correct_answer = str(row[5]).strip().upper() if len(row) > 5 and row[5] else ''
                
                for i, letter in enumerate(option_letters):
                    option_text = str(row[i + 1]).strip() if len(row) > i + 1 and row[i + 1] else ''
                    if option_text:
                        options.append({
                            'id': f"{question_id}_opt_{i}",
                            'text': option_text,
                            'is_correct': letter == correct_answer
                        })
                
                if question_text and options:
                    questions.append({
                        'id': question_id,
                        'text': question_text,
                        'options': options
                    })
            
            return questions
            
        except Exception as e:
            raise ValueError(f"Failed to parse Excel file: {str(e)}")
    
    def parse_text(self, text_content: str) -> List[dict]:
        """
        Parse plain text content for questions
        Useful for copy-paste scenarios
        """
        questions = []
        current_question = None
        current_options = []
        
        lines = text_content.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Check for question
            question_match = self.QUESTION_PATTERN.match(line)
            if question_match:
                if current_question and current_options:
                    questions.append(self._create_question_from_text(
                        current_question, current_options
                    ))
                
                current_question = question_match.group(2).strip()
                current_options = []
                continue
            
            # Check for option
            option_match = self.OPTION_PATTERN.match(line)
            if option_match and current_question:
                # Check for bold markers like **text** or __text__
                option_text = option_match.group(2).strip()
                is_bold = '**' in line or '__' in line
                
                # Clean markdown bold markers
                option_text = re.sub(r'\*\*(.+?)\*\*', r'\1', option_text)
                option_text = re.sub(r'__(.+?)__', r'\1', option_text)
                
                current_options.append({
                    'letter': option_match.group(1).upper(),
                    'text': option_text,
                    'is_correct': is_bold
                })
        
        # Last question
        if current_question and current_options:
            questions.append(self._create_question_from_text(
                current_question, current_options
            ))
        
        return questions
    
    def _create_question_from_text(self, question_text: str, options: List[dict]) -> dict:
        """Create question dict from text parsing"""
        question_id = str(uuid.uuid4())[:8]
        
        formatted_options = []
        for i, opt in enumerate(options):
            formatted_options.append({
                'id': f"{question_id}_opt_{i}",
                'text': opt['text'],
                'is_correct': opt['is_correct']
            })
        
        return {
            'id': question_id,
            'text': question_text,
            'options': formatted_options
        }


# Singleton instance
parser = QuestionParser()


def parse_document(file_content: bytes, filename: str) -> Tuple[List[dict], Optional[str]]:
    """
    Main entry point for parsing documents
    
    Args:
        file_content: Binary content of the file
        filename: Original filename to determine type
        
    Returns:
        Tuple of (questions list, error message or None)
    """
    try:
        if filename.lower().endswith('.docx'):
            questions = parser.parse_docx(file_content)
        elif filename.lower().endswith(('.xlsx', '.xls')):
            questions = parser.parse_excel(file_content)
        elif filename.lower().endswith('.txt'):
            questions = parser.parse_text(file_content.decode('utf-8'))
        else:
            return [], f"Unsupported file format: {filename}"
        
        return questions, None
        
    except Exception as e:
        return [], str(e)
